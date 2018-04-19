import openpyxl


#To load the xlsx spread sheet using object creation
wb=openpyxl.load_workbook('dataset.xlsx')

print(wb)

#Used to get list of sheet name from a excel document
##sheet= wb.sheetnames
##print(sheet)

#This will return the sheet object,we can use this sheet
#object to access the sheet value
##sheet=wb.get_sheet_by_name('Sheet1')
##print(sheet)


#Accessin the sheet cell value
##cell=sheet['A2'].value
##print(cell)

#Used for viewing all the sheets available in a particular workbook
##print(wb.get_sheet_names())

#Checking What happens when i acces the datatime in excel
##sheet=wb.get_sheet_by_name('Sheet1')
##print(sheet['A10'].value)

#Accessing multiple cells in excel using for loop
##sheet=wb.get_sheet_by_name('Sheet1')
##for i in range(1,10):
##   print(sheet['A'+str(i)].value)

#Printing the whole sheet and accessin all cells using for loop
##sheet=wb.get_sheet_by_name('Sheet1')
##for i in range(1,10):
##   print(sheet['A'+str(i)].value,sheet['B'+str(i)].value,sheet['C'+str(i)].value,sheet['D'+str(i)].value,sheet['E'+str(i)].value,sheet['F'+str(i)].value,sheet['G'+str(i)].value,sheet['H'+str(i)].value)

#Using cell method we can access the values in a cell by passin the row and column value as parameters
##sheet=wb.get_sheet_by_name('Sheet1')
##print(sheet.cell(row=1,column=1).value)

#using for retrieving all value from the specified row and column value
##sheet=wb.get_sheet_by_name('Sheet1')
##for i in range(1,10):
##   print(sheet.cell(row=i,column=2).value)








   
