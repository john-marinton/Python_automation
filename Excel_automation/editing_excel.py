import openpyxl


#Creating excel workbook by using .Workbook method and creating object to access it 
wb=openpyxl.Workbook()
print(wb)

#used for creating new sheet inside a workbook
wb.create_sheet()
print(wb.get_sheet_names())

#creating an new sheet inside workbook at a specifed position
wb.create_sheet(index=0,title='my own sheet')

#Adding values to an empty excel sheet
sheet=wb.get_sheet_by_name('Sheet')
sheet['A1']='Hello'
print(sheet['A1'].value)


#changin the title of an sheet
sheet=wb.get_sheet_by_name('Sheet')
sheet.title='Edited sheet'


#Saving the created excel sheet
wb.save('edited.xlsx')
